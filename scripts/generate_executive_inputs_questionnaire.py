#!/usr/bin/env python3
"""Generate Executive Inputs questionnaire: one worksheet per deck tile."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent.parent / "docs" / "Executive_Inputs_Questionnaire.xlsx"


def style_header(ws, row=1):
    for cell in ws[row]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def autosize(ws, max_w=78):
    for col in ws.iter_cols(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        letter = get_column_letter(col[0].column)
        w = min(max_w, max(len(str(c.value or "")) for c in col) + 2)
        ws.column_dimensions[letter].width = w


def add_questions(ws, questions: list[str]):
    ws.append(["#", "Question"])
    style_header(ws)
    for i, q in enumerate(questions, start=1):
        ws.append([i, q])
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    autosize(ws)


def main():
    wb = Workbook()
    wb.remove(wb.active)

    # Tile 1: Catalog & economics
    ws1 = wb.create_sheet("Catalog economics", 0)
    add_questions(
        ws1,
        [
            "What program or course types will you offer at launch (e.g. degree, diploma, certificate, coaching, short-term skill)?",
            "For each program type at launch, what is the exact list of fields that must appear on the public listing page?",
            "For each listing field, is it mandatory or optional at publish time?",
            "Who is the named approver (role and person) for publishing or updating a live program listing?",
            "What is the default currency for all fees shown to applicants?",
            "Are listed program fees inclusive of GST, exclusive of GST, or mixed by category? Specify per category if mixed.",
            "If GST applies, what is the GST rate or HSN/SAC code (or mapping rule) for each fee type or program category?",
            "Is the applicant fee a single full amount, or are multiple fee components shown separately (e.g. tuition vs registration)? List each component.",
            "If instalments are allowed, what is the exact schedule (number of instalments, due dates, and any interest or late fee rule)?",
            "If only full payment is allowed, confirm that no partial payment state should exist in the product at launch.",
            "What is the application fee or booking fee amount per program, or what rule determines it when it varies?",
            "Who is authorised to create, edit, or retire a program listing—the platform operator, each institute, or both?",
            "If institutes edit listings, what approval step (if any) must occur before changes go live?",
            "What revenue share or commission applies to each institute (percentage of which base: gross, net of gateway fees, or other)?",
            "Is revenue share identical for all institutes or defined per partner? If per partner, where is the authoritative schedule maintained?",
            "What minimum contractual or onboarding artefacts must exist before an institute’s programs can appear (e.g. signed agreement, KYC pack)?",
            "What eligibility or quality criteria must an institute meet before onboarding?",
            "What is the rule and responsible party when a program is cancelled, rescheduled, or intake is closed after applications exist?",
            "Who decides and executes refunds when a program is cancelled by the institute vs by the platform?",
            "Do you offer bundles, cross-institute packages, or waitlists at launch? Describe each rule.",
            "What catalog change cadence do you expect (e.g. annual intake only vs rolling admissions), and who communicates cut-off dates to students?",
        ],
    )

    # Tile 2: Brand & policy
    ws2 = wb.create_sheet("Brand policy")
    add_questions(
        ws2,
        [
            "What is the legal registered name of the entity that will operate the platform and sign contracts?",
            "What customer-facing brand name must appear in the product if it differs from the legal name?",
            "Have you supplied master logo files in vector format (e.g. SVG, EPS, AI) and raster formats (e.g. PNG) with clear space rules?",
            "What are the official primary and secondary brand colours (hex or Pantone)?",
            "What is the official public support email address?",
            "What is the official postal address for legal notices and consumer grievances?",
            "Who will provide final Terms of Use text, and in which version date should engineering implement?",
            "Who will provide final Privacy Policy text, and in which version date should engineering implement?",
            "Will you provide a Cookie Policy or is it embedded within the Privacy Policy only?",
            "What jurisdictions or laws should be named in governing-law clauses (e.g. India, specific state)?",
            "What is the standard full-refund window, in calendar days, measured from which event (payment success, batch start, first class)?",
            "List each scenario in which a partial refund is allowed and the formula or percentage for each.",
            "List each scenario in which no refund is allowed after payment.",
            "Who holds decision authority for discretionary refunds (role), and what is the target decision turnaround (hours/days)?",
            "Are any refund rules mandated by a university, board, or statute that override your default policy? Attach or summarise.",
            "What exact disclaimer or non-refundable language must appear on the application or checkout step?",
            "What languages must the public-facing copy support at launch (e.g. English only, English + Hindi)?",
            "Do you require WCAG or other accessibility standards for marketing pages and application flows? Specify level if yes.",
            "Who approves final UI copy for emails and SMS templates (role)?",
        ],
    )

    # Tile 3: Payments
    ws3 = wb.create_sheet("Payments")
    add_questions(
        ws3,
        [
            "Which payment gateway provider is approved for production (exact product name, e.g. Razorpay Standard, Razorpay Route)?",
            "Will you use a single legal merchant account or a marketplace / split-settlement model? State which.",
            "What is the exact registered legal name that must appear on the gateway merchant profile and receipts?",
            "What is the legal entity type (e.g. Pvt Ltd, LLP, trust) for that merchant?",
            "Who is the signatory completing gateway KYC and accepting the merchant agreement?",
            "Will you provide separate sandbox (test) API credentials and production API credentials?",
            "Who will receive and securely store the production key_id and key_secret, and in which secret manager or vault?",
            "Who is authorised to rotate gateway credentials, and how will engineering be notified without exposing secrets in email?",
            "Which payment methods must be enabled at launch: UPI, domestic debit card, domestic credit card, netbanking, wallets, EMI, Pay Later?",
            "Must international cards be accepted at launch?",
            "If international cards are used, what currencies must be supported for display and settlement?",
            "What is the maximum single transaction amount you expect at launch (for gateway limit and risk review)?",
            "What is the expected typical transaction amount range (min and max) during the first admission cycle?",
            "Will you use payment links only, hosted checkout, server-side order creation, or a combination? Specify.",
            "If using webhooks, what is the production webhook URL owner (team) and who monitors webhook failures?",
            "What is the required idempotency policy for creating orders or payment attempts (who retries, when)?",
            "Into which bank account(s) should settled funds ultimately arrive? Provide one row per account: account holder legal name, bank name, IFSC, account type.",
            "If funds are split between parties, what exact percentage or fixed rupee amount goes to each beneficiary for a standard ₹1,00,000 fee payment?",
            "Is the split calculated on the gross amount paid by the student or on the net amount after gateway fees? State the rule.",
            "Who bears the gateway transaction fee (MDR) and GST on that fee—the platform, the institute, the student (surcharge), or a defined split?",
            "If MDR is shared, what is the exact allocation rule (e.g. platform 100%, 50/50, deducted from institute share only)?",
            "Do you require instant settlement, standard T+N settlement, or hybrid? State N if standard.",
            "Who reconciles settlements against internal ledgers (role), and how often (daily, weekly)?",
            "What is the process when a payment succeeds but the application record fails—who refunds or corrects, and within what SLA?",
            "What is the process for chargebacks, disputes, or “request for information” from the bank or gateway?",
            "Must the system issue tax invoices or receipts to students? If yes, who is the issuer of supply on the invoice?",
            "What GSTIN (if applicable) must appear on student-facing invoices?",
            "What invoice numbering series and prefix rules must finance enforce?",
            "Are partial capture, recurring payments, or subscriptions in scope for launch?",
            "If refunds are initiated from the dashboard, who approves each refund (role), and must partial refunds be supported?",
            "What student-visible reference (order id, payment id, receipt number) must appear on confirmation emails?",
            "Does PCI scope require that card data never touch your servers (confirm hosted fields / redirect-only)?",
        ],
    )

    # Tile 4: Operating model
    ws4 = wb.create_sheet("Operating model")
    add_questions(
        ws4,
        [
            "Who is the single primary liaison for product, scope, and launch decisions (full name, role, email, phone)?",
            "Who is the designated backup liaison with the same authority if the primary is unavailable?",
            "What is your definition of a “lead” for reporting (e.g. form started, form submitted, call logged, paid application)?",
            "Which system is the system of record for leads (name of CRM or spreadsheet), and should the platform push events there?",
            "If CRM integration is required, what API keys or webhooks can you provide, and who administers that system?",
            "Which channels will you use for applicant support at launch: email, phone, WhatsApp, in-app ticket, other? List all.",
            "What are the staffed hours and timezone for each support channel?",
            "What is the target first-response time for each channel during staffed hours?",
            "What is the target resolution time for payment-related tickets vs general information tickets?",
            "Who triages tickets that involve both an institute and the platform?",
            "Who notifies students of application status changes (platform team, institute, or automated only)?",
            "Who receives operational alerts for failed payments, webhook errors, or integration outages (names or mailing lists)?",
            "How should institutes access applicant or payment summaries (portal role, CSV export, email digest)? Specify frequency.",
            "What SLA do institutes expect from the platform operator for data corrections or listing updates?",
            "Who owns escalation to legal when a student threatens litigation or regulatory complaint?",
            "Who coordinates communications during a production incident (name/role)?",
        ],
    )

    # Tile 5: Compliance
    ws5 = wb.create_sheet("Compliance")
    add_questions(
        ws5,
        [
            "List every category of personal data you will collect at launch (e.g. name, phone, email, DOB, address, government ID number, marksheets, photograph).",
            "For each data category, what is the specific purpose of collection (e.g. admission, invoicing, marketing)?",
            "What lawful basis do you assert for processing each category (consent, contract, legal obligation, legitimate interest)?",
            "How long will you retain personal data for an active user account after last login or last transaction?",
            "How long will you retain personal data for a closed or deleted account after erasure request (if any grace period)?",
            "How long will you retain payment and settlement records for audit and tax purposes?",
            "How long will you retain marketing consent records and proof of opt-in?",
            "Will marketing emails/SMS require a separate explicit opt-in from transactional messages?",
            "What is the default state of any marketing checkbox at signup (unchecked is recommended—confirm your choice)?",
            "Provide the exact opt-in wording you require for marketing communications.",
            "What age defines a “minor” on your platform (e.g. under 18 years as of which date)?",
            "For minors, what parental or guardian consent is required before account creation or payment?",
            "Which product features (if any) must be blocked or limited for minor accounts until consent is verified?",
            "Who is the designated grievance or data-protection contact (name, email, postal address)?",
            "What is the committed response time for privacy or data-access requests?",
            "Will personal data be stored or processed on servers outside India? If yes, list countries and safeguards.",
            "Will you share personal data with institutes, payment gateways, SMS/email vendors, or analytics tools? List each recipient category.",
            "For each recipient category, what data fields are shared and for what purpose?",
            "Do you require a DPIA, student data policy, or state-specific education rules to be reflected in product behaviour?",
            "Must the product support data export and erasure workflows? If yes, who approves erasure that conflicts with legal retention?",
        ],
    )

    # Tile 6: Infrastructure
    ws6 = wb.create_sheet("Infrastructure")
    add_questions(
        ws6,
        [
            "What exact production hostname(s) and apex domain will users type to reach the application?",
            "Who owns the domain registrar account, and can they grant DNS edit access to the engineering team?",
            "If DNS is managed elsewhere (e.g. Cloudflare), who holds that account and can create API tokens?",
            "Are there corporate policies restricting DNS TTL values, CAA records, or prohibited DNS providers?",
            "Will production workloads run in your cloud account, the vendor’s account, or a shared account? Specify cloud vendor if known.",
            "In which region or availability zones must application data reside at rest (city or region name)?",
            "Is data residency restricted to India only for application databases and file storage?",
            "Who provisions TLS certificates, and are they required to be from a specific CA?",
            "What is the target go-live date and time (with timezone)?",
            "Are there blackout periods when production changes are forbidden (e.g. admission peaks)? List each window.",
            "What recurring maintenance window is acceptable for deployments (weekday/weekend, start time, end time, timezone)?",
            "Who has authority to approve production releases (name/role) and who approves emergency hotfixes?",
            "Do you require a dedicated staging environment that mirrors production integrations? Who hosts it?",
            "What uptime or availability target should the architecture aim for (e.g. 99.5%, 99.9%)?",
            "Who receives on-call alerts for application errors, downtime, or security findings (team, email, PagerDuty, etc.)?",
            "Are VPNs, IP allowlists, or MDM required for staff accessing production consoles?",
            "What log retention period is required in production (days/months) for application and access logs?",
        ],
    )

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
