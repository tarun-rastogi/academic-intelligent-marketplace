#!/usr/bin/env python3
"""Business / company information to collect from Merrakii for the student-facing app (not technical build specs)."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent.parent / "docs" / "Merrakii_Business_Capture.xlsx"


def style_header(ws, row=1):
    for cell in ws[row]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def autosize(ws, max_w=88):
    for col in ws.iter_cols(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        letter = get_column_letter(col[0].column)
        w = min(max_w, max(len(str(c.value or "")) for c in col) + 2)
        ws.column_dimensions[letter].width = w


def add_sheet(wb, name: str, intro: str, questions: list[str]):
    """Column A: question, column B: answer (empty until submit)."""
    ws = wb.create_sheet(name)
    ws.append(["Question", "Answer"])
    style_header(ws)
    if intro:
        ws.append(["—", intro])
        for cell in ws[2]:
            cell.font = Font(italic=True)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for q in questions:
        ws.append([q, ""])
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    autosize(ws)


def add_submission_log(wb, index: int = 0):
    """Submission log: one row per Section + Q#; resubmits update the same row (server-side upsert)."""
    ws = wb.create_sheet("Submission log", index)
    ws.append(["Timestamp (UTC)", "Section", "Q#", "Question", "Response"])
    style_header(ws)
    autosize(ws)


def main():
    wb = Workbook()
    wb.remove(wb.active)
    add_submission_log(wb, 0)

    add_sheet(
        wb,
        "Company content",
        "Copy and facts Merrakii must supply for public pages (you decide layout and UX).",
        [
            "What short tagline or one-line value proposition should appear with the Merrakii brand on the home or landing experience?",
            "What ‘About Merrakii’ narrative should students read (mission, years in business, geography served)? Prefer word count or bullet limit if you have one.",
            "What differentiators should we emphasise versus generic ed-tech (e.g. offline office, counsellor-led, exam prep + admissions)?",
            "List any factual claims we may display (e.g. number of students counselled, partner institutes, success stories)—and confirm each is accurate and approved for marketing.",
            "May we name specific institutes or exams as ‘partners’ or ‘we help with’ on the site? If yes, list exact wording Merrakii approves; if no, say so.",
            "Do you want an ‘Our team’ or counsellor profile section? If yes, provide names, roles, photos, and bios—or confirm ‘no’.",
            "Do you want to show office photos or a virtual tour link? Provide assets or URLs.",
            "Do you have student testimonials Merrakii authorises for the app? Provide quote text, first name or initials, and written permission where required.",
            "What languages should customer-facing marketing copy use at launch (e.g. English only, English + Hindi)?",
            "Any words or phrases Merrakii wants avoided in all copy (legal, brand, or competitive reasons)?",
        ],
    )

    add_sheet(
        wb,
        "Brand and contact",
        "Identity and every ‘reach us’ touchpoint students see.",
        [
            "What is the legal registered name of the entity that owns the business (for receipts and legal footers)?",
            "What customer-facing brand name must appear if it differs from the legal name?",
            "Provide logo files (master formats you have) and primary/secondary brand colours for the app shell.",
            "What is the main office address to show (street, city, state, PIN)—for ‘Visit us’ or contact page?",
            "What is the main public email for inquiries (e.g. hello@…)?",
            "What phone number(s) should appear for general inquiries? Label each (e.g. admissions, billing) if different.",
            "What are Merrakii’s official website and social profile URLs (Instagram, LinkedIn, YouTube, etc.) to link from the app?",
            "What are regular business hours for the office and for phone support (timezone included)?",
            "Should we display a map link (Google Maps) for the office? If yes, provide the URL or coordinates.",
        ],
    )

    add_sheet(
        wb,
        "Human support",
        "When a student needs a person: numbers, channels, rules.",
        [
            "When a student needs human help during search or application, what is the primary phone number to call?",
            "Is that number the same for post-enrollment questions (fees, documents) or different? List each number and purpose.",
            "Are calls toll-free, standard STD, or WhatsApp voice? Specify.",
            "What hours is live phone support staffed? What should we tell users outside those hours (e.g. leave SMS, callback)?",
            "Do you offer WhatsApp chat for support? If yes, provide the business number and any welcome message Merrakii wants.",
            "Do you offer email ticketing only for certain topics? List topic → email address.",
            "Do you want an in-app ‘Request a counsellor callback’ flow? If yes, what fields may we collect (name, phone, preferred time)?",
            "What is the maximum callback commitment Merrakii wants displayed (e.g. within 24 business hours)?",
            "For escalations (payment failed, document rejected), who is the internal role name students hear (e.g. ‘admissions desk’)—not personal names unless required?",
            "Should we show the physical office as an option (‘Visit us for in-person counselling’)? If yes, confirm address and whether appointment is required.",
            "Any channel Merrakii does not want in the app (e.g. no Telegram)?",
        ],
    )

    add_sheet(
        wb,
        "Payment gateway",
        "Merrakii’s vendor and contract drive what we must wire in the product; collect facts, not engineering lectures.",
        [
            "Which payment gateway company is Merrakii contracted with (legal product name)?",
            "Who at Merrakii is the signatory or owner relationship with that gateway (name, role, email) for escalations?",
            "Who is the gateway’s technical or account manager contact (name, email) Merrakii can introduce to your team?",
            "Will Merrakii provide sandbox credentials for integration testing and production credentials before go-live?",
            "Does Merrakii’s contract include split or route settlements to institutes, or does Merrakii settle manually after receiving funds? Describe the intended money flow in one paragraph.",
            "If splits are in the gateway: what percentage or amount goes to Merrakii vs each institute category (or attach Merrakii’s schedule)?",
            "Which payment methods must be offered at checkout (UPI, debit, credit, netbanking, wallets, EMI, Pay Later)?",
            "Are international cards or foreign currency in scope? Yes or no.",
            "What fee amount does the student see before pay—full programme fee, registration fee only, or variable by institute? Who approves the fee line items shown?",
            "Who bears the gateway transaction fee (MDR)—Merrakii, the institute, the student via surcharge, or mixed?",
            "What refund and cancellation policy must the checkout and help content state (full refund window, partial, non-refundable cases)? Provide Merrakii-approved wording or bullet rules.",
            "Who at Merrakii approves refunds in practice (role), and typical turnaround Merrakii commits to for students?",
            "What must appear on the payment receipt or tax invoice to the student (legal name, GSTIN, address, HSN/SAC if applicable)?",
            "Does Merrakii require EMI or ‘pay in parts’ messaging even if not enabled at launch?",
            "Any institute-specific payment rules we must display on certain courses?",
        ],
    )

    add_sheet(
        wb,
        "Legal and consent",
        "Merrakii’s lawyer or leadership supplies or approves text; you implement flows.",
        [
            "Who will provide final Terms of Use (or Terms & Conditions) text for the application—Merrakii legal, external counsel, or approve a draft you supply?",
            "Who will provide final Privacy Policy text, including how student data is used and shared with institutes?",
            "Should cookie or similar tracking consent be shown? If yes, does Merrakii have approved wording?",
            "For phone number and OTP login: what consent line must the user accept (exact or approved paraphrase)?",
            "For browser or device location to suggest nearby institutes: what consent explanation must we show, and may we proceed if the user denies location?",
            "For push notifications (application status, payment): what opt-in wording does Merrakii require?",
            "For sharing the student’s application data with a chosen institute: what consent must we capture (checkbox text)?",
            "For marketing messages (new courses, webinars): separate opt-in required? Default on or off per Merrakii policy?",
            "For document uploads (ID, marksheets): what declaration must the user confirm (authenticity, purpose)?",
            "Governing law and jurisdiction clause Merrakii requires in footer or legal screens (e.g. courts of which city)?",
            "Registered grievance or nodal officer name, email, and postal address for consumer complaints (if Merrakii must display this under policy or practice).",
            "Any mandatory disclaimers for counselling vs guaranteed admission (non-guarantee language Merrakii wants everywhere)?",
        ],
    )

    add_sheet(
        wb,
        "Programs and partners",
        "Business data Merrakii owns: what the app can list and sell.",
        [
            "What types of offerings does Merrakii place online at launch (institute admissions only, exam coaching, both, career counselling packages)?",
            "Who provides the authoritative list of institutes, courses, exams, and fees that students can browse—Merrakii spreadsheet, CRM export, or manual updates by Merrakii staff?",
            "How often does Merrakii expect prices or intake dates to change, and who is responsible to notify your team for app updates?",
            "Which institutes or boards has Merrakii formally authorised to represent for online applications through this app?",
            "Are there categories Merrakii does not want listed online (e.g. certain exams or geographies)?",
            "What happens if a course is full or delisted after a student started an application—who communicates with the student and what is Merrakii’s policy text?",
            "For the student dashboard: what labels does Merrakii want for enrolment states (e.g. Applied, Under review, Confirmed, Rejected)?",
            "What fee fields must the dashboard show (amount paid via app, balance due, total programme fee)—confirm definitions Merrakii uses with students.",
            "Does Merrakii need institute-specific terms or addenda shown before apply on certain listings?",
        ],
    )

    add_sheet(
        wb,
        "Operations handoff",
        "Who does what after launch (business ops, not your engineering SLA).",
        [
            "Who is Merrakii’s single product owner for content and policy changes on the live app (name, email)?",
            "Who receives alerts when a student completes payment or uploads documents (role or shared inbox)?",
            "Who confirms ‘enrolment confirmed’ with the institute—Merrakii only, institute only, or joint—and should the app reflect only Merrakii-confirmed status?",
            "What SLA does Merrakii want to promise students for application review (e.g. 2 business days)—if any?",
        ],
    )

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
