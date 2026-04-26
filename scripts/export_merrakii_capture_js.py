#!/usr/bin/env python3
"""Regenerate docs/merrakii-business-capture-data.js from Merrakii_Business_Capture.xlsx."""

import json
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "docs" / "Merrakii_Business_Capture.xlsx"
OUT = ROOT / "docs" / "merrakii-business-capture-data.js"


def main():
    wb = load_workbook(XLSX, read_only=True, data_only=True)
    out: dict[str, dict[str, str | list[str]]] = {}
    for name in wb.sheetnames:
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        intro = ""
        questions: list[str] = []
        start = 1
        # Row 1: headers (Question, Answer). Optional row 2: "—" + intro. Questions in column A.
        if len(rows) > 1 and str(rows[1][0] or "").strip() == "—":
            intro = str(rows[1][1] or "").strip()
            start = 2
        for r in rows[start:]:
            if not r or not r[0]:
                continue
            q = str(r[0]).strip()
            if q and q != "—":
                questions.append(q)
        out[name] = {"intro": intro, "questions": questions}

    OUT.write_text(
        "window.MERRAKII_BUSINESS_CAPTURE = " + json.dumps(out, ensure_ascii=False) + ";\n",
        encoding="utf-8",
    )
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
