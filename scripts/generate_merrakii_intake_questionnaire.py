#!/usr/bin/env python3
"""Merrakii client intake: questions for building the marketplace app, split by urgency."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent.parent / "docs" / "Merrakii_Intake_Questionnaire.xlsx"


def style_header(ws, row=1):
    for cell in ws[row]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def autosize(ws, max_w=85):
    for col in ws.iter_cols(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        letter = get_column_letter(col[0].column)
        w = min(max_w, max(len(str(c.value or "")) for c in col) + 2)
        ws.column_dimensions[letter].width = w


def add_sheet(wb, name: str, questions: list[str]):
    ws = wb.create_sheet(name)
    ws.append(["#", "Question"])
    style_header(ws)
    for i, q in enumerate(questions, start=1):
        ws.append([i, q])
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    autosize(ws)


def main():
    wb = Workbook()
    wb.remove(wb.active)

    needed_now = [
        "What is the production domain name (and optional staging subdomain) we should target for Day-1 hosting on AWS?",
        "Who will create the AWS account root / billing owner—Merrakii or the build team—and who grants IAM access to deploy frontend, API, and database?",
        "What are the legal entity name and customer-facing brand name to display across web and mobile?",
        "Provide master logo assets (vector + PNG) and primary/secondary brand colours (hex) for web and app shells.",
        "What is the public support email and phone number shown in the app and on policies?",
        "Confirm the end-to-end user journey you want at MVP: browser location → phone → OTP → AI search → catalogue with filters → course → apply → personal details → Aadhaar/10th/12th upload → pay. List any step you want added, removed, or reordered.",
        "For browser geolocation: should we use coarse location only, prompt every visit, or remember consent for the session? Any restriction on showing location to users who deny permission?",
        "What country dialling code(s) are in scope for phone login (India +91 only or others)?",
        "Will OTP be SMS-only at go-live, or also WhatsApp/voice fallback? (Integration is on our side; confirm Merrakii-approved user-facing copy for the OTP message.)",
        "For India SMS DLT: will Merrakii register the principal entity and sender ID, or is another party the registered entity? Provide sender ID / PE name as applicable for template filing.",
        "Provide the definitive list of user intents the AI search must handle at MVP (e.g. discover by city, by stream, by fee band, by exam, compare institutes)—and the expected in-app action for each intent.",
        "Provide 15–20 example natural-language queries students might type, with the expected result (filters applied or institute/course list).",
        "List every filter dimension on the institute catalogue at MVP (e.g. state, city, course category, fee range, accreditation, mode online/offline). For each, specify allowed values or whether it is free text.",
        "Should catalogue default sort be relevance, distance from user location, fee low-to-high, or configurable? What is the default for MVP?",
        "What fields define a “course” card on the catalogue and on the course detail page before Apply (title, duration, eligibility, fees, start date, seats, syllabus link, etc.)?",
        "Who supplies the initial institute and course data for build/UAT—Merrakii spreadsheet, institute portal, or API—and what is the minimum number of institutes/courses required to freeze UI?",
        "For Apply: besides full name, address, and age, list every additional mandatory and optional field (email, gender, category, parent name, etc.).",
        "Are address fields structured (pin code → auto city/state) or free text? Which country constraints apply?",
        "Confirm the exact list of uploads at MVP: Aadhaar, class 10 mark sheet, class 12 mark sheet—any others (photo, signature, category certificate)?",
        "Per document type: allowed file formats (PDF/JPG/PNG), maximum file size, and whether masking last 4 digits of Aadhaar is required in UI guidance.",
        "Who is allowed to view or download uploaded documents in Merrakii’s operating model (platform admin only, institute users, both)?",
        "What payment gateway vendor is Merrakii contracted with (product name, e.g. Razorpay Standard vs Route)?",
        "Provide the technical integration pack from that vendor: API documentation URL, sandbox dashboard access or sandbox key_id/key_secret, and a named technical SPOC at the gateway vendor.",
        "Describe the settlement model we must implement: single merchant settlement vs split/route to multiple linked accounts; percentage or fixed split between Merrakii and each institute.",
        "For each settlement beneficiary (Merrakii platform share and each institute pattern): legal name as on bank, and whether the gateway uses linked account IDs vs manual settlement outside the gateway.",
        "Which payment methods must be enabled at checkout for go-live (UPI, cards, netbanking, wallets, EMI, Pay Later)?",
        "Must international cards or multi-currency display be supported at MVP?",
        "Who bears the gateway MDR and GST on fees, and should the student see an all-inclusive fee or line-item gateway surcharge?",
        "What student-facing receipt or invoice data must appear after payment (GSTIN, legal name, HSN/SAC if any)?",
        "What are the refund and cancellation rules that the checkout and post-payment flows must enforce in copy and workflow?",
        "Confirm push notifications are required on Day 1 for web (if applicable), Android, and iOS—or phased: which channel first?",
        "For push: what categories of events need notifications at MVP (application submitted, payment success, document rejected, offer letter)?",
        "Will Merrakii supply Firebase/APNs configuration later only, or is a Merrakii-owned Firebase/Apple Developer account required from week one?",
        "What peak concurrent users and requests per minute should we size for at Month-1 go-live (you indicated ~100k users; confirm burst vs sustained)?",
        "Confirm environment strategy: separate dev/staging/prod AWS accounts or single account with VPC separation?",
        "Who owns draft Terms of Use, Privacy Policy, and cookie/consent text for the first deploy—Merrakii legal or external counsel—and who signs off before public staging?",
    ]

    wait_15 = [
        "Provide production API credentials for the payment gateway (key_id/key_secret or equivalent) and any webhook signing secret—after sandbox sign-off.",
        "Confirm production webhook URL(s) and whether the gateway requires IP allowlisting of outbound API calls from AWS.",
        "Final production merchant KYC status and any go-live checklist item from the gateway vendor (e.g. settlement account verification complete).",
        "Invite the build team to Merrakii’s Apple Developer Program (Admin or App Manager role) and confirm Team ID for signing iOS builds.",
        "Invite the build team to Google Play Console (Release Manager) for Android upload signing and internal testing tracks.",
        "Provide Apple Push Notification service key (.p8), Key ID, Team ID, and bundle identifiers for dev/staging/prod—or confirm Merrakii will create them and share securely.",
        "Provide Firebase project (or FCM server key) for Android push in production, and confirm google-services.json delivery process.",
        "Final App Store and Play Store listing: app name, subtitle, short/long description, keywords, category, support URL, marketing URL.",
        "Provide store screenshots and feature graphic specifications if Merrakii supplies creative, or approve placeholders.",
        "Complete App Store privacy nutrition label answers and Play Data safety form based on actual data collection in the shipped MVP.",
        "Final production institute and course catalogue export (full list) if MVP used a subset during build.",
        "Confirm any institute-specific terms or additional uploads required after MVP (e.g. entrance score, interview slot)—to schedule phase 2.",
        "Production SMS template IDs (DLT) after approval, if different from sandbox or if templates change post-review.",
        "Final signed-off legal pages (PDF or HTML) and version date string to embed in the app footer.",
        "If Merrakii requires a specific analytics or crash-reporting dashboard (e.g. their own GA4 property), provide property ID and data-retention preferences.",
        "Confirm data-retention period for application documents and PII after rejection or withdrawal, for production cron/jobs.",
        "Any enterprise MDM or SSO requirement for Merrakii staff admin portals (if in scope)—identity provider metadata.",
        "Certificate pinning or extra security audit requirements for production release.",
        "Disaster recovery expectations: RPO/RTO for database and file storage backups on AWS.",
        "Named operational contacts for production incidents (24/7 or business hours) and escalation matrix.",
    ]

    add_sheet(wb, "Needed immediately", needed_now)
    add_sheet(wb, "Can wait 15 days", wait_15)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
