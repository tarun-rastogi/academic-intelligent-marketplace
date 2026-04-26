#!/usr/bin/env python3
"""Generate India higher-education reference workbook (4 sheets). Upload .xlsx to Google Drive → Open with Google Sheets."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent.parent / "docs" / "India_Education_Catalog.xlsx"


def style_header(ws, row=1):
    for cell in ws[row]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def autosize_columns(ws, max_width=52):
    for col in ws.iter_cols(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        letter = get_column_letter(col[0].column)
        width = min(
            max_width,
            max(len(str(c.value or "")) for c in col) + 2,
        )
        ws.column_dimensions[letter].width = width


def main():
    wb = Workbook()

    # --- Sheet 1: Fields of study ---
    ws1 = wb.active
    ws1.title = "Fields of Study"
    ws1.append(
        [
            "Broad domain",
            "Field of study",
            "Typical specialisations / sub-areas",
            "Common delivery formats in India",
        ]
    )
    fields = [
        (
            "Natural sciences",
            "Physics",
            "Theoretical, condensed matter, optics, nuclear, particle, plasma, biophysics",
            "UG/PG degrees, integrated MSc, research programmes, online NPTEL/SWAYAM",
        ),
        (
            "Natural sciences",
            "Chemistry",
            "Organic, inorganic, physical, analytical, environmental, materials",
            "UG/PG, PhD, lab-intensive diplomas at polytechnics for applied chemistry",
        ),
        (
            "Natural sciences",
            "Mathematics & statistics",
            "Pure maths, applied maths, actuarial science, biostatistics, data statistics",
            "UG/PG, BStat/MStat (ISI pathway), online degrees and certificates",
        ),
        (
            "Natural sciences",
            "Earth & environmental sciences",
            "Geology, geophysics, meteorology, oceanography, environmental science",
            "UG/PG, field programmes, IGNOU/ODL options for some streams",
        ),
        (
            "Life sciences",
            "Biological sciences",
            "Botany, zoology, microbiology, genetics, ecology, marine biology",
            "UG/PG, integrated programmes, research",
        ),
        (
            "Life sciences",
            "Biotechnology & biochemistry",
            "Industrial biotech, medical biotech, enzymology, molecular biology",
            "UG/PG, BTech Biotech, skill diplomas",
        ),
        (
            "Agriculture & allied",
            "Agriculture, horticulture, forestry, fisheries",
            "Agronomy, plant breeding, soil science, dairy, poultry, agribusiness",
            "BSc/MSc Agri, BVSc, diplomas, ICAR-accredited programmes, distance via SAUs",
        ),
        (
            "Engineering & technology",
            "Core engineering",
            "Civil, mechanical, electrical, electronics, chemical, metallurgy, mining, textile, ceramic, production, automobile, marine, aerospace, agricultural engg., food technology",
            "BE/BTech, ME/MTech, diplomas, part-time and working-professional programmes",
        ),
        (
            "Computing & IT",
            "Computer science & software engineering",
            "Algorithms, systems, networks, compilers, HCI, software engineering",
            "BCA, BSc CS, BTech CSE/IT, MCA, MSc CS, online UG/PG, bootcamps",
        ),
        (
            "Computing & IT",
            "Information systems & cybersecurity",
            "Information security, digital forensics, cloud, enterprise systems",
            "Degrees, PG diplomas, vendor certifications, short-term ODL",
        ),
        (
            "Artificial intelligence & data",
            "AI, machine learning & robotics",
            "ML, deep learning, NLP, computer vision, robotics, autonomous systems",
            "BTech/MTech AI, MSc Data Science, PG diplomas, industry micro-credentials",
        ),
        (
            "Artificial intelligence & data",
            "Data science & analytics",
            "Big data, business analytics, econometrics, operations research",
            "UG/PG, MBA analytics, online specialisations, executive certificates",
        ),
        (
            "Medicine & health sciences",
            "Allopathic medicine (MBBS pathway)",
            "Pre-clinical, para-clinical, clinical departments; public health orientation",
            "MBBS, MD/MS, DM/MCh, diplomas (DGO, DCH, etc.), fellowship training",
        ),
        (
            "Medicine & health sciences",
            "Dentistry, AYUSH, nursing, pharmacy",
            "BDS/MDS; Ayurveda, Homeopathy, Unani, Siddha; BSc nursing, GNM; BPharm/MPharm",
            "Degrees regulated by NMC/DCI/NCISM/NCH/Nursing councils/PCI",
        ),
        (
            "Medicine & health sciences",
            "Allied health & rehabilitation",
            "Physiotherapy, occupational therapy, optometry, MLT, radiography, audiology, BASLP",
            "UG/PG degrees and diplomas; clinical practicum-heavy formats",
        ),
        (
            "Economics, commerce & management",
            "Economics",
            "Micro, macro, econometrics, development, international, agricultural economics",
            "BA/BSc Economics, MA/MSc, PhD, MBA economics electives",
        ),
        (
            "Economics, commerce & management",
            "Commerce, accounting & finance",
            "Accounting, taxation, auditing, banking, insurance, CFA-oriented coursework",
            "BCom, MCom, BBA, MBA, professional CA/CS/CMA parallel coaching",
        ),
        (
            "Economics, commerce & management",
            "Management & business",
            "HR, marketing, operations, strategy, entrepreneurship, family business",
            "BBA, MBA, PGDM, executive programmes, online MBAs",
        ),
        (
            "Social sciences & humanities",
            "Sociology, political science, international relations, social work",
            "Public policy, governance, gender studies, development studies",
            "BA, MA, MSW, MPP, PhD, ODL from open universities",
        ),
        (
            "Social sciences & humanities",
            "History, philosophy, languages & literature",
            "Indology, comparative literature, linguistics, translation studies",
            "BA, MA, MPhil, PhD, certificate courses in languages",
        ),
        (
            "Social sciences & humanities",
            "Psychology & cognitive science",
            "Clinical, counselling, organisational, cognitive neuroscience interfaces",
            "BA/BSc Psych, MA/MSc, MPhil Clinical Psych (RCI-regulated pathways)",
        ),
        (
            "Law & governance",
            "Legal studies",
            "Constitutional, criminal, corporate, IP, human rights, environmental law",
            "BA LLB, BBA LLB, LLB, LLM, PhD, diploma in labour laws etc.",
        ),
        (
            "Education & sports",
            "Education & teacher preparation",
            "Curriculum, pedagogy, EdTech, special education, adult education",
            "BEd, MEd, DElEd, BElEd, physical education degrees",
        ),
        (
            "Architecture, planning & design",
            "Architecture & planning",
            "Urban design, landscape, conservation, regional planning",
            "BArch, MArch, BPlan, MPlan; COA norms for internships",
        ),
        (
            "Architecture, planning & design",
            "Design, fine arts & performing arts",
            "Product, communication, fashion, interior, animation, music, dance, drama",
            "BDes, MDes, BFA, MFA, BMus; NID/NIFT-style portfolios",
        ),
        (
            "Media, communication & liberal arts",
            "Journalism, film, mass communication",
            "Digital media, documentary, broadcast, new media law/ethics",
            "BA, MA, MJMC, PG diplomas, film schools",
        ),
        (
            "Hospitality, tourism & culinary",
            "Hotel, tourism & culinary arts",
            "F&B operations, housekeeping, travel, event management",
            "BHM, BTTM, diplomas, NCHM JEE-linked programmes",
        ),
        (
            "Veterinary & animal sciences",
            "Veterinary science & animal husbandry",
            "Surgery, medicine, pathology, livestock production",
            "BVSc & AH, MVSc, diplomas in dairy/veterinary fields",
        ),
        (
            "Library, information & documentation",
            "Library & information science",
            "Digital libraries, knowledge management, archives",
            "BLISc, MLISc, certificates",
        ),
        (
            "Defence & strategic studies",
            "Defence, strategic & security studies",
            "Geopolitics, defence technology policy, NCC-linked academic tracks",
            "UG/PG electives, specialised university departments, NDA/CDS preparation",
        ),
        (
            "Interdisciplinary & emerging",
            "Interdisciplinary & NEP-aligned clusters",
            "Environmental studies, sustainability, climate, innovation, public health policy",
            "4-year UG with research, multidisciplinary minors, twinning/honours",
        ),
    ]
    for row in fields:
        ws1.append(list(row))
    style_header(ws1)
    autosize_columns(ws1)

    # --- Sheet 2: Degree courses ---
    ws2 = wb.create_sheet("Degree Courses")
    ws2.append(
        [
            "Qualification / degree title",
            "Level (India)",
            "Typical nominal duration",
            "Broad stream",
            "Regulator / quality notes (indicative)",
        ]
    )
    degrees = [
        ("Certificate (credit-bearing UGC)", "Below bachelor", "1 year or less", "Various", "HEI under UGC framework; not equivalent to full degree"),
        ("Diploma (Polytechnic / engineering)", "Sub-degree", "3 years (10+2 entry) or 2 years (lateral)", "Engineering & technology", "AICTE/state board of tech. education"),
        ("Diploma (Pharmacy DPharm)", "Sub-degree", "2 years", "Pharmacy", "PCI"),
        ("Diploma in Elementary Education (DElEd)", "Sub-degree", "2 years", "Education", "NCTE"),
        ("Bachelor of Arts (BA)", "Undergraduate", "3 years (4-year honours in many universities)", "Humanities & social sciences", "University/conventional UG"),
        ("Bachelor of Science (BSc)", "Undergraduate", "3 years (4-year honours)", "Sciences", "University"),
        ("Bachelor of Commerce (BCom)", "Undergraduate", "3 years", "Commerce", "University"),
        ("Bachelor of Business Administration (BBA)", "Undergraduate", "3 years", "Management", "University/AICTE-linked institutes"),
        ("Bachelor of Computer Applications (BCA)", "Undergraduate", "3 years", "Computing", "University"),
        ("Bachelor of Engineering / Technology (BE/BTech)", "Undergraduate", "4 years", "Engineering & technology", "AICTE + university"),
        ("Bachelor of Architecture (BArch)", "Undergraduate", "5 years", "Architecture", "COA"),
        ("Bachelor of Planning (BPlan)", "Undergraduate", "4 years", "Planning", "University/institute of planning"),
        ("Bachelor of Pharmacy (BPharm)", "Undergraduate", "4 years", "Pharmacy", "PCI"),
        ("Bachelor of Science (Nursing) / PB BSc Nursing", "Undergraduate", "4 years / 2 years", "Nursing", "INC"),
        ("Bachelor of Physiotherapy (BPT)", "Undergraduate", "4–4.5 years", "Allied health", "State councils / university norms"),
        ("Bachelor of Occupational Therapy (BOT)", "Undergraduate", "4–4.5 years", "Allied health", "University/council norms"),
        ("Bachelor of Optometry (BOptom)", "Undergraduate", "4 years", "Allied health", "University"),
        ("Bachelor of Audiology & Speech-Language Pathology (BASLP)", "Undergraduate", "4 years", "Allied health", "RCI-related programmes"),
        ("MBBS", "Undergraduate professional", "5.5 years incl. internship", "Medicine (allopathic)", "NMC"),
        ("BDS", "Undergraduate professional", "5 years incl. internship", "Dentistry", "DCI"),
        ("BAMS / BHMS / BUMS / BSMS / BYNS", "Undergraduate professional", "5.5 years incl. internship", "AYUSH systems", "NCISM / state boards"),
        ("BVSc & AH", "Undergraduate professional", "5.5 years", "Veterinary", "VCI"),
        ("LLB (3-year)", "Undergraduate professional", "3 years", "Law", "BCI (after UG)"),
        ("BA LLB / BBA LLB / BSc LLB (integrated)", "Undergraduate professional", "5 years", "Law", "BCI"),
        ("BEd", "Undergraduate professional", "2 years (after UG) / integrated variants", "Education", "NCTE"),
        ("Bachelor of Physical Education (BPEd / B.P.Ed)", "Undergraduate", "3–4 years", "Sports & education", "NCTE/sports authority norms"),
        ("Bachelor of Fine Arts (BFA)", "Undergraduate", "4 years", "Fine arts", "University"),
        ("Bachelor of Design (BDes)", "Undergraduate", "4 years", "Design", "Institute-specific"),
        ("Bachelor of Hotel Management (BHM)", "Undergraduate", "3–4 years", "Hospitality", "University/AICTE"),
        ("Bachelor of Tourism & Travel Management (BTTM)", "Undergraduate", "3–4 years", "Tourism", "University"),
        ("Bachelor of Library & Information Science (BLISc)", "Undergraduate", "1 year (after UG)", "Library science", "University"),
        ("Bachelor of Social Work (BSW)", "Undergraduate", "3 years", "Social work", "University"),
        ("Integrated BSc–MSc / BS–MS", "Undergraduate + PG integrated", "5 years", "Sciences", "IISER/IISc-like pathways; universities"),
        ("Integrated BTech–MTech", "Undergraduate + PG integrated", "5 years", "Engineering", "Some IITs/NITs/IIITs"),
        ("Master of Arts (MA)", "Postgraduate", "2 years", "Humanities & social sciences", "University"),
        ("Master of Science (MSc)", "Postgraduate", "2 years", "Sciences", "University"),
        ("Master of Commerce (MCom)", "Postgraduate", "2 years", "Commerce", "University"),
        ("Master of Business Administration (MBA)", "Postgraduate", "2 years (1-year executive variants)", "Management", "UGC/AICTE norms; PGDM by AICTE"),
        ("Post Graduate Diploma in Management (PGDM)", "Postgraduate diploma", "2 years", "Management", "AICTE-approved autonomous B-schools"),
        ("Master of Computer Applications (MCA)", "Postgraduate", "2 years", "Computing", "UGC"),
        ("Master of Engineering / Technology (ME/MTech)", "Postgraduate", "2 years", "Engineering", "AICTE + GATE common entry"),
        ("Master of Architecture (MArch)", "Postgraduate", "2 years", "Architecture", "COA"),
        ("Master of Planning (MPlan)", "Postgraduate", "2 years", "Planning", "University"),
        ("Master of Pharmacy (MPharm)", "Postgraduate", "2 years", "Pharmacy", "PCI"),
        ("Master of Laws (LLM)", "Postgraduate", "1–2 years", "Law", "BCI"),
        ("MEd / MA Education", "Postgraduate", "2 years", "Education", "NCTE"),
        ("Master of Social Work (MSW)", "Postgraduate", "2 years", "Social work", "University"),
        ("Master of Fine Arts (MFA) / Master of Design (MDes)", "Postgraduate", "2 years", "Arts & design", "University/institute"),
        ("MD / MS (Medicine)", "Postgraduate medical", "3 years", "Clinical specialties", "NMC"),
        ("MDS", "Postgraduate dental", "3 years", "Dental specialties", "DCI"),
        ("MD/MS (AYUSH)", "Postgraduate", "3 years", "AYUSH", "NCISM"),
        ("DM / MCh", "Doctorate-level super-specialty", "3 years (after MD/MS)", "Super-specialties", "NMC"),
        ("MVSc", "Postgraduate", "2 years", "Veterinary", "VCI"),
        ("MPhil", "Postgraduate research", "1–2 years (being phased down)", "Various", "UGC historical framework; check current regulations"),
        ("PhD / Doctor of Philosophy", "Doctoral", "3–5+ years", "All major streams", "UGC regulations; NET/JRF common"),
        ("Doctor of Medicine (DM) alternative naming — avoid confusion with UK MD", "—", "—", "—", "In India MD is PG medical, not UG"),
        ("Executive MBA / PGPEX", "Postgraduate executive", "1 year", "Management", "IIMs and peers"),
        ("IPM (Integrated Programme in Management)", "UG+PG management", "5 years", "Management", "IIM Indore/Rohtal etc."),
    ]
    for row in degrees:
        ws2.append(list(row))
    style_header(ws2)
    autosize_columns(ws2)

    # --- Sheet 3: Entrance exams ---
    ws3 = wb.create_sheet("Entrance Exams")
    ws3.append(
        [
            "Examination",
            "Primary level / purpose",
            "Domain / stream",
            "Conducting body (indicative)",
            "Typical use in India",
        ]
    )
    exams = [
        ("JEE Main", "UG engineering", "Engineering & architecture (Paper 2)", "NTA", "NITs, IIITs, GFTIs, state/private admissions, JEE Advanced filter"),
        ("JEE Advanced", "UG engineering", "IITs", "IITs (rotating)", "IIT UG programmes"),
        ("NEET-UG", "UG medical & related", "MBBS, BDS, AYUSH UG, BVSc, BSc Nursing (as notified)", "NTA", "All-India quota + state admissions per MCC/state rules"),
        ("NEET-PG", "PG medical", "MD/MS/Diploma (allopathic)", "NTA", "PG medical counselling"),
        ("NEET-SS", "Super-specialty", "DM/MCh seats", "NBE", "Super-specialty entrance"),
        ("INI-CET / INI-SS", "PG & super-specialty", "AIIMS, JIPMER, NIMHANS, etc.", "AIIMS", "Institute-level PG/SS"),
        ("NEET-MDS", "PG dental", "MDS", "NBE/NTA (check year notification)", "Dental PG"),
        ("CUET-UG", "UG central universities", "Multiple UG programmes", "NTA", "Central and participating universities"),
        ("CUET-PG", "PG central universities", "Multiple PG programmes", "NTA", "Central and participating universities"),
        ("GATE", "PG engineering + jobs", "ME/MTech/PhD; PSUs", "IISc + IITs on rotation", "M Tech admissions, scholarships, PSU recruitment"),
        ("CAT", "PG management", "MBA/PGP (IIMs and others)", "IIMs", "Flagship MBA entrance"),
        ("XAT", "PG management", "XLRI & associates", "XLRI", "B-school admissions"),
        ("SNAP", "PG management", "Symbiosis institutes", "SIU", "Symbiosis PG programmes"),
        ("NMAT", "PG management", "NMIMS & partners", "GMAC delivery", "NMIMS and allied"),
        ("CMAT", "PG management", "AICTE-approved programmes", "NTA", "MBA/PGDM admissions"),
        ("MAT / ATMA", "PG management", "Various private B-schools", "AIMA / AIMS", "Seasonal MBA admissions"),
        ("IIFT MBA (IB) entrance", "PG management", "International business", "NTA", "IIFT MBA"),
        ("CLAT", "UG & PG law", "NLUs and others", "Consortium of NLUs", "5-year LLB, LLM"),
        ("AILET", "UG & PG law", "NLU Delhi", "NLU Delhi", "Law programmes"),
        ("LSAT—India", "UG/PG law", "Partner law schools", "Pearson VUE / LSAC", "Multiple private NLU+ schools"),
        ("SLAT", "UG law", "Symbiosis law schools", "SIU", "Symbiosis law"),
        ("BITSAT", "UG engineering", "BITS Pilani campuses", "BITS", "BITS UG"),
        ("VITEEE", "UG engineering", "VIT", "VIT", "VIT admissions"),
        ("SRMJEEE", "UG engineering", "SRM", "SRMIST", "SRM admissions"),
        ("MET (Manipal)", "UG health/engineering/etc.", "Manipal Academy", "MAHE", "Manipal UG/PG as applicable"),
        ("WBJEE", "UG engineering/pharm/arch", "West Bengal", "WBJEEB", "State quota institutions"),
        ("MHT-CET", "UG engineering/pharm/agri", "Maharashtra", "State CET cell", "State admissions"),
        ("KCET", "UG engineering/pharm/agri", "Karnataka", "KEA", "State admissions"),
        ("COMEDK UGET", "UG engineering", "Karnataka private unaided", "COMEDK", "Private engineering seats"),
        ("AP EAPCET (EAMCET)", "UG engg/pharm/agri", "Andhra Pradesh", "APSCHE", "State admissions"),
        ("TS EAMCET", "UG engg/pharm/agri", "Telangana", "TSCHE", "State admissions"),
        ("KEAM", "UG engg/medical/arch", "Kerala", "CEE Kerala", "State admissions"),
        ("Goa CET", "UG engg/pharm/nursing", "Goa", "DTE Goa", "State admissions"),
        ("OJEE", "UG/PG Odisha", "Multiple", "OJEE board", "State admissions"),
        ("UPSEE (legacy; check current UP state CET)", "State", "Uttar Pradesh", "AKTU/state", "Verify current UP engineering/pharm rules"),
        ("IPU CET", "UG/PG", "GGSIPU programmes", "GGSIPU", "Delhi NCR admissions"),
        ("NIFT Entrance", "UG/PG design", "Fashion design", "NIFT", "NIFT programmes"),
        ("NID DAT", "UG/PG design", "Industrial design", "NID", "NID programmes"),
        ("UCEED", "UG design", "BDes at IITs", "IIT Bombay", "IIT BDes"),
        ("CEED", "PG design", "MDes at IITs/IISc", "IIT", "MDes admissions"),
        ("NATA", "UG architecture", "BArch (alongside JEE Main Paper 2)", "Council of Architecture", "Architecture aptitude"),
        ("NCHM JEE", "UG hospitality", "Hotel management", "NTA", "IHM admissions"),
        ("ICAR AIEEA UG/PG", "Agriculture UG/PG", "Agricultural sciences", "NTA", "ICAR institutes & seats"),
        ("UGC-NET", "Assistant Professor / JRF", "All UGC NET subjects", "NTA", "Eligibility for college teaching & JRF"),
        ("CSIR-UGC NET", "JRF / lectureship", "Science subjects", "NTA", "Research fellowship eligibility"),
        ("ICMR-JRF / DBT-JRF etc.", "Research fellowship", "Life sciences", "Respective agencies", "PhD support schemes"),
        ("GPAT", "PG pharmacy", "MPharm", "NTA", "MPharm + scholarships"),
        ("CEED alternative: NID PG", "PG design", "Design", "NID", "See NID notifications"),
        ("JNUEE / JNU-specific", "UG/PG", "Humanities & sciences", "JNU", "JNU admissions (pattern changes—verify)"),
        ("DUET", "UG/PG", "Delhi University", "NTA", "DU entrance programmes"),
        ("BHU UET/PET", "UG/PG", "Banaras Hindu University", "NTA", "BHU programmes"),
        ("AIIMS BSc Nursing / paramedical entrances", "UG allied health", "Nursing/paramedical", "AIIMS", "Institute-specific"),
        ("Indian Navy / Army / Air Force recruitment exams", "Officer/other ranks", "Defence", "CDS, AFCAT, INET, etc.", "Commissioned officer entry"),
        ("NDA & NA Examination", "10+2 entry defence", "Army/Navy/Air Force", "UPSC", "NDA training"),
        ("CDS Examination", "Graduate defence", "OTA/IMA/INA/AFA", "UPSC", "Officer entry after graduation"),
        ("UPSC Civil Services (CSE)", "All-India services", "IAS/IPS/IFS etc.", "UPSC", "Premier administrative recruitment"),
        ("UPSC CAPF (AC)", "Paramilitary", "Assistant Commandant", "UPSC", "CAPF leadership"),
        ("UPSC Engineering Services (ESE)", "Technical services", "Railways, CPWD, etc.", "UPSC", "Class-A technical posts"),
        ("State PSC examinations", "State services", "Administrative & technical", "State PSCs", "State civil services"),
        ("SSC CGL / CHSL / JE etc.", "Central govt. jobs", "Non-gazetted/gazetted per exam", "SSC", "Staff selection"),
        ("RRB NTPC / JE / ALP etc.", "Railways", "Technical & non-technical", "RRBs", "Railway recruitment"),
        ("IBPS PO/Clerk/SO", "Banking", "PSU banks", "IBPS", "Public sector bank recruitment"),
        ("SBI PO/Clerk", "Banking", "State Bank", "SBI", "SBI recruitment"),
        ("RBI Grade B / Assistant", "Banking", "Reserve Bank", "RBI", "RBI recruitment"),
        ("CA Foundation/Intermediate/Final", "Professional accounting", "Chartered Accountancy", "ICAI", "CA qualification entry"),
        ("CS Executive/Professional", "Corporate law", "Company Secretary", "ICSI", "CS qualification"),
        ("CMA Foundation/Intermediate/Final", "Cost & management accountancy", "CMA", "ICMAI", "CMA qualification"),
        ("Actuarial exams (IAI)", "Actuarial science", "Insurance & risk", "IAI", "Actuary certification"),
        ("CFA Program (global)", "Investment analysis", "Finance", "CFA Institute", "Taken by Indian candidates for finance careers"),
        ("FRM / PRM etc.", "Risk management", "Finance", "GARP / PRMIA", "Professional risk credentials"),
        ("TOEFL / IELTS / PTE", "English proficiency", "Study abroad", "ETS/British Council/Pearson", "Foreign admissions from India"),
        ("GRE / GMAT", "PG abroad / some India programmes", "Various", "ETS/GMAC", "Overseas PG; some Indian exec programmes"),
        ("SAT / ACT", "UG abroad", "Undergraduate abroad", "College Board / ACT", "US-style UG applications"),
        ("TIFR GS / JEST / JAM", "Research PG", "Physics/maths/chem/biology", "TIFR/IITs etc.", "Research institute admissions"),
        ("CEED / IIT JAM / JEST", "MSc integrated/PhD pathways", "Sciences", "Various", "Premier science admissions"),
        ("Army Welfare Education Society tests", "Schools / colleges", "AWES", "AWES", "Army education system"),
        ("Hotel management institute-specific", "Diploma/UG", "Hospitality", "Private chains", "Direct institute tests"),
        ("State law CETs (e.g. MH CET Law)", "UG law", "State universities", "State authorities", "Non-NLU law colleges"),
        ("Design institute-specific (Pearl, UID, etc.)", "UG/PG design", "Private design", "Institutes", "Portfolio + test"),
        ("Film & television institute entrances (FTII, SRFTI, etc.)", "PG diploma / courses", "Film", "Institutes", "Creative arts"),
    ]
    for row in exams:
        ws3.append(list(row))
    style_header(ws3)
    autosize_columns(ws3)

    # --- Sheet 4: Non-degree courses ---
    ws4 = wb.create_sheet("Non-Degree Courses")
    ws4.append(
        [
            "Category",
            "Examples / schemes / formats",
            "Typical providers",
            "Duration & credential (indicative)",
        ]
    )
    nondeg = [
        ("Industrial Training Institute (ITI) trades", "Electrician, fitter, welder, COPA, mechanic trades", "NCVT/SCVT-affiliated ITIs", "6 months–2 years; NCVT certificate"),
        ("National Skill Qualification Framework (NSQF) aligned short courses", "Sector skill councils’ QP-aligned programmes", "NSDC partners, training partners", "Variable; certificate"),
        ("Pradhan Mantri Kaushal Vikas Yojana (PMKVY)", "Fee-based and RPL certifications", "PMKVY TCs", "Short-term; monetary reward in some schemes"),
        ("Advanced Diploma / PG Certificate (non-degree)", "PG diploma in corporate law, clinical research, etc.", "Universities, autonomous institutes", "1 year; not always equivalent to master’s degree"),
        ("Certificate programmes (university)", "Certificate in Tally, GST, data analytics, French, etc.", "Universities, IGNOU, online HEIs", "3–12 months"),
        ("SWAYAM / NPTEL certificates", "MOOCs with proctored exam", "MHRD/AICTE/NPTEL", "Semester-length modules"),
        ("IGNOU / State Open University certificates", "Awareness, vocational, awareness in law, yoga", "ODL universities", "6 months–1 year"),
        ("NIOS (Open School)", "Senior secondary, vocational subjects", "NIOS", "Flexible pacing; board-level certificate"),
        ("Coaching & test preparation", "JEE, NEET, CUET, CLAT, UPSC, CA/CS/CMA, banking, SSC", "Private coaching industry", "3 months–2+ years; no formal degree"),
        ("Executive certificates & open programmes", "IIM/IIT executive education, ISB certificates", "Premier institutes", "Few days–6 months"),
        (
            "Project management contact programmes (institute)",
            "Certificate/diploma in project management; PMP exam-prep bootcamps; CAPM prep; PRINCE2 Foundation/Practitioner; Agile/Scrum (PSM/CSM-style) prep; MS Project / scheduling labs",
            "Private training chains (e.g. NIIT, Aptech), ed-tech partners, PMI R.E.P.s, management institutes",
            "Typically 3–6 months; institute completion certificate; PMP/PRINCE2/PMI credentials are awarded by the governing body after passing their exam (and meeting eligibility where required)",
        ),
        (
            "ITIL & IT service management short courses",
            "ITIL 4 Foundation and follow-on modules; ITSM practice workshops; sometimes bundled with IT operations career tracks",
            "Axelos/PeopleCert authorised training orgs, global and Indian training partners, corporate L&D vendors",
            "Often 2–5 days to a few weeks for Foundation; longer blended programmes up to ~3–6 months when packaged with internships or multiple modules; exam-based global certification",
        ),
        (
            "Python & programming-language institute certificates",
            "Python certification / diploma programmes; Java, C++, data structures; full-stack web (Django/Flask, MERN); SQL & databases; often job-oriented ‘6-month’ software training",
            "Private IT training institutes, university extension centres, bootcamp brands",
            "Commonly 3–6 months (also 2–9 month variants); institute certificate; sometimes mapped to vendor exams (e.g. Oracle Java) optionally",
        ),
        (
            "Artificial Intelligence & machine learning (short institute programmes)",
            "Applied AI/ML, deep learning, NLP, computer vision, GenAI/LLM application tracks; MLOps introductions; often marketed as PG diploma/certificate (non-UGC degree)",
            "Ed-tech platforms, private institutes, some university centres of continuing education",
            "Typically 3–6 months (intensive) to 9–12 months for deeper diplomas; portfolio/capstone; credential is usually institute or platform certificate unless paired with university credit",
        ),
        (
            "Professional frameworks: Six Sigma, risk, quality (institute format)",
            "Lean Six Sigma (Yellow/Green/Black belt contact classes); ISO internal auditor; IT governance/compliance short programmes",
            "Consulting firms, QCI-linked training bodies, institutes",
            "3–6 month belts common for Green Belt; exam + project components vary by body",
        ),
        ("Vendor / industry certifications", "AWS, Azure, GCP, Cisco CCNA, Oracle, Red Hat, Salesforce, Kubernetes (CKA/CKAD)", "Vendors + authorised training partners", "Exam-based credentials; prep courses often 4–12 weeks or bundled in 3–6 month career programmes"),
        ("Digital marketing & UX bootcamps", "Performance marketing, SEO, product design sprints", "Private academies", "8–24 weeks"),
        ("Yoga & wellness certificates", "YCB levels, wellness diplomas", "AYUSH Ministry recognised bodies", "Variable"),
        ("Beauty, wellness & aviation hospitality diplomas", "Cosmetology, cabin crew, ground staff", "Private institutes", "6–18 months"),
        ("Foreign language diplomas (non-degree)", "DELF, DELE, Goethe, JLPT prep programmes", "Institutes & alliances", "Level-based"),
        ("Financial markets certifications", "NISM modules, NCFM (legacy)", "SEBI/NSE", "Module-wise"),
        ("Teacher refresher / in-service", "CBSE/NCERT workshops, DIET programmes", "Government & boards", "Short cycles"),
        ("Apprenticeship embedded training", "NAPS-linked on-the-job learning", "Industry + MSDE ecosystem", "1–2 years stipend-based"),
        ("Micro-credentials & credit banks (emerging)", "Stackable credits under ABC/NEP pilots", "HEIs + DigiLocker", "Variable"),
        ("Community college / bridge programmes", "Foundational numeracy, bridge to UG", "State pilots, NGOs", "Short term"),
        ("Hospital-based diplomas (paramedical)", "OTT, anesthesia assistant, radiography diploma", "Hospitals + councils", "2–3 years; registration rules vary by state"),
        ("GNM / ANM (nursing diplomas)", "General nursing, auxiliary nurse midwife", "Nursing schools", "3 years / 2 years; INC norms evolving—verify current rules"),
        ("Maritime & shipping courses", "GP rating, pre-sea courses", "DG Shipping-approved institutes", "STCW-aligned"),
        ("Fire & safety diplomas", "Industrial safety, fire tech", "State boards / private", "1–2 years"),
        ("Agriculture extension short programmes", "KVK trainings, farmer FPO skilling", "ICAR-KVK, NGOs", "Days to weeks"),
    ]
    for row in nondeg:
        ws4.append(list(row))
    style_header(ws4)
    autosize_columns(ws4)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
