"""
Serve docs/digi.html and POST /api/executive-inputs to merge form answers into Merrakii_Business_Capture.xlsx.

Run from repo root: python -m executive-inputs-server.app
Or: cd executive-inputs-server && python app.py
"""

from __future__ import annotations

import os
from pathlib import Path

from flask import Flask, jsonify, request, send_from_directory
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
DOCS = ROOT / "docs"
XLSX = DOCS / "Merrakii_Business_Capture.xlsx"
SUBMIT_TOKEN = os.environ.get("EXECUTIVE_INPUTS_SUBMIT_TOKEN", "").strip()

app = Flask(__name__)


def _question_data_rows(ws):
    """Return 1-based Excel row numbers for each question row (column A, after header + optional intro)."""
    max_r = ws.max_row or 1
    intro_offset = 0
    v2 = ws.cell(row=2, column=1).value
    if v2 is not None and str(v2).strip() == "—":
        intro_offset = 1
    out = []
    for r in range(2 + intro_offset, max_r + 1):
        ac = ws.cell(row=r, column=1).value
        if ac is None:
            continue
        t = str(ac).strip()
        if not t or t == "—":
            continue
        out.append(r)
    return out


def _apply_answers(xlsx_path: Path, payload: dict) -> None:
    if not xlsx_path.is_file():
        raise FileNotFoundError(f"Workbook not found: {xlsx_path}")

    wb = load_workbook(xlsx_path, read_only=False, data_only=False)
    for tab_name, answers in payload.items():
        if not isinstance(answers, dict) or tab_name not in wb.sheetnames:
            continue
        ws = wb[tab_name]
        qrows = _question_data_rows(ws)
        for qi_s, value in answers.items():
            try:
                qi = int(qi_s)
            except (TypeError, ValueError):
                continue
            if qi < 0 or qi >= len(qrows):
                continue
            row = qrows[qi]
            ws.cell(row=row, column=2, value="" if value is None else str(value))
    wb.save(xlsx_path)


@app.after_request
def _cors(resp):
    # Allow same-repo static usage and local file opened via http server
    resp.headers["Access-Control-Allow-Origin"] = os.environ.get("CORS_ALLOW_ORIGIN", "*")
    resp.headers["Access-Control-Allow-Methods"] = "GET, POST, OPTIONS"
    resp.headers["Access-Control-Allow-Headers"] = "Content-Type, X-Submit-Token"
    return resp


@app.route("/api/executive-inputs", methods=["OPTIONS"])
def options_exec():
    return ("", 204)


@app.post("/api/executive-inputs")
def post_exec():
    if SUBMIT_TOKEN:
        if request.headers.get("X-Submit-Token", "") != SUBMIT_TOKEN:
            return jsonify({"ok": False, "error": "unauthorized"}), 401
    if not request.is_json:
        return jsonify({"ok": False, "error": "expected application/json"}), 400
    data = request.get_json()
    if not isinstance(data, dict):
        return jsonify({"ok": False, "error": "body must be a JSON object"}), 400
    path = Path(os.environ.get("EXECUTIVE_INPUTS_XLSX", str(XLSX)))
    try:
        _apply_answers(path, data)
    except FileNotFoundError as e:
        return jsonify({"ok": False, "error": str(e)}), 500
    return jsonify({"ok": True, "path": str(path)})


@app.get("/")
def index():
    return send_from_directory(DOCS, "digi.html")


@app.get("/<path:filename>")
def static_docs(filename: str):
    if ".." in filename or filename.startswith("/"):
        return ("Not found", 404)
    target = (DOCS / filename).resolve()
    try:
        target.relative_to(DOCS.resolve())
    except ValueError:
        return ("Not found", 404)
    if not target.is_file():
        return ("Not found", 404)
    return send_from_directory(DOCS, filename)


def main():
    port = int(os.environ.get("PORT", "5050"))
    # threaded=True for development; use gunicorn in production
    app.run(host=os.environ.get("HOST", "127.0.0.1"), port=port, debug=False, threaded=True)


if __name__ == "__main__":
    main()
